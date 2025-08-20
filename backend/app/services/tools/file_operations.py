try:
    from openai.agents import tool
except Exception:  # pragma: no cover
    def tool(*args, **kwargs):
        def decorator(func):
            return func
        return decorator

import os
import json
import csv
from pathlib import Path


@tool(name="read_file", description="Read the contents of a file from the file system.")
def read_file(file_path: str, encoding: str = "utf-8"):
    """Read and return the contents of a file."""
    try:
        if not os.path.exists(file_path):
            return {"error": f"File not found: {file_path}"}
        
        with open(file_path, 'r', encoding=encoding) as file:
            content = file.read()
        
        return {
            "content": content,
            "file_path": file_path,
            "size": len(content),
            "attachments": [
                {"type": "text", "content": f"File: {file_path}\nSize: {len(content)} characters\n\nContent:\n{content}"}
            ]
        }
    except Exception as e:
        return {"error": f"Error reading file: {str(e)}"}


@tool(name="write_file", description="Write content to a file on the file system.")
def write_file(file_path: str, content: str, encoding: str = "utf-8"):
    """Write content to a file, creating directories if needed."""
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'w', encoding=encoding) as file:
            file.write(content)
        
        return {
            "success": True,
            "file_path": file_path,
            "bytes_written": len(content.encode(encoding)),
            "attachments": [
                {"type": "text", "content": f"Successfully wrote {len(content)} characters to {file_path}"}
            ]
        }
    except Exception as e:
        return {"error": f"Error writing file: {str(e)}"}


@tool(name="list_directory", description="List files and directories in a given path.")
def list_directory(directory_path: str, include_hidden: bool = False):
    """List contents of a directory with file details."""
    try:
        if not os.path.exists(directory_path):
            return {"error": f"Directory not found: {directory_path}"}
        
        items = []
        for item in os.listdir(directory_path):
            if not include_hidden and item.startswith('.'):
                continue
            
            item_path = os.path.join(directory_path, item)
            stat = os.stat(item_path)
            
            items.append({
                "name": item,
                "path": item_path,
                "type": "directory" if os.path.isdir(item_path) else "file",
                "size": stat.st_size,
                "modified": stat.st_mtime
            })
        
        # Sort by type (directories first), then by name
        items.sort(key=lambda x: (x["type"] != "directory", x["name"]))
        
        return {
            "directory": directory_path,
            "items": items,
            "count": len(items),
            "attachments": [
                {"type": "text", "content": f"Directory listing for {directory_path}:\n" + 
                 "\n".join([f"{'📁' if item['type'] == 'directory' else '📄'} {item['name']}" for item in items[:20]])}
            ]
        }
    except Exception as e:
        return {"error": f"Error listing directory: {str(e)}"}


@tool(name="parse_csv", description="Parse a CSV file and return structured data.")
def parse_csv(file_path: str, delimiter: str = ",", has_header: bool = True):
    """Parse CSV file and return structured data."""
    try:
        if not os.path.exists(file_path):
            return {"error": f"CSV file not found: {file_path}"}
        
        data = []
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter=delimiter)
            
            if has_header:
                headers = next(reader)
                for row in reader:
                    data.append(dict(zip(headers, row)))
            else:
                for i, row in enumerate(reader):
                    data.append({"row": i, "values": row})
        
        return {
            "data": data,
            "row_count": len(data),
            "file_path": file_path,
            "attachments": [
                {"type": "text", "content": f"Parsed CSV: {file_path}\nRows: {len(data)}\nSample: {json.dumps(data[:3], indent=2) if data else 'No data'}"}
            ]
        }
    except Exception as e:
        return {"error": f"Error parsing CSV: {str(e)}"}


@tool(name="create_csv", description="Create a CSV file from structured data.")
def create_csv(file_path: str, data: list, headers: list = None):
    """Create CSV file from list of dictionaries or lists."""
    try:
        # Determine the final file path - use downloads directory for public access
        downloads_dir = Path(__file__).parent.parent.parent.parent / "public" / "downloads"
        downloads_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate a unique filename if only name provided, otherwise use the provided path
        if "/" not in file_path and "\\" not in file_path:
            # Just a filename, save to downloads directory
            filename = file_path if file_path.endswith('.csv') else f"{file_path}.csv"
            # Add timestamp to make it unique
            import time
            timestamp = int(time.time())
            base_name = filename.rsplit('.csv', 1)[0]
            final_filename = f"{base_name}_{timestamp}.csv"
            final_file_path = downloads_dir / final_filename
        else:
            # Full path provided, but also copy to downloads for URL access
            final_file_path = Path(file_path)
            os.makedirs(final_file_path.parent, exist_ok=True)
            
            # Also create a copy in downloads directory for URL access
            filename = final_file_path.name
            import time
            timestamp = int(time.time())
            base_name = filename.rsplit('.csv', 1)[0]
            download_filename = f"{base_name}_{timestamp}.csv"
            download_file_path = downloads_dir / download_filename
        
        # Write CSV data
        def write_csv_data(file_path):
            with open(file_path, 'w', newline='', encoding='utf-8') as file:
                if isinstance(data[0], dict):
                    # Data is list of dictionaries
                    fieldnames = headers or list(data[0].keys())
                    writer = csv.DictWriter(file, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
                else:
                    # Data is list of lists
                    writer = csv.writer(file)
                    if headers:
                        writer.writerow(headers)
                    writer.writerows(data)
        
        # Save the file
        write_csv_data(str(final_file_path))
        
        # If we created a separate download copy, save that too
        if 'download_file_path' in locals():
            write_csv_data(str(download_file_path))
            download_url = f"http://localhost:8000/files/download/{download_filename}"
            public_filename = download_filename
        else:
            download_url = f"http://localhost:8000/files/download/{final_file_path.name}"
            public_filename = final_file_path.name
        
        return {
            "success": True,
            "file_path": str(final_file_path),
            "download_url": download_url,
            "public_filename": public_filename,
            "rows_written": len(data),
            "attachments": [
                {"type": "text", "content": f"✅ Created CSV file: {public_filename} with {len(data)} rows\n\n📥 Download link: {download_url}"}
            ]
        }
    except Exception as e:
        return {"error": f"Error creating CSV: {str(e)}"}


@tool(name="create_excel", description="Create an Excel (.xlsx) file from structured data.")
def create_excel(file_path: str, data: list, sheet_name: str = "Sheet1", headers: list = None):
    """Create Excel file from list of dictionaries or lists."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import uuid
        
        # Determine the final file path - use downloads directory for public access
        downloads_dir = Path(__file__).parent.parent.parent.parent / "public" / "downloads"
        downloads_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate a unique filename if only name provided, otherwise use the provided path
        if "/" not in file_path and "\\" not in file_path:
            # Just a filename, save to downloads directory
            filename = file_path if file_path.endswith('.xlsx') else f"{file_path}.xlsx"
            # Add timestamp to make it unique
            import time
            timestamp = int(time.time())
            base_name = filename.rsplit('.xlsx', 1)[0]
            final_filename = f"{base_name}_{timestamp}.xlsx"
            final_file_path = downloads_dir / final_filename
        else:
            # Full path provided, but also copy to downloads for URL access
            final_file_path = Path(file_path)
            os.makedirs(final_file_path.parent, exist_ok=True)
            
            # Also create a copy in downloads directory for URL access
            filename = final_file_path.name
            import time
            timestamp = int(time.time())
            base_name = filename.rsplit('.xlsx', 1)[0]
            download_filename = f"{base_name}_{timestamp}.xlsx"
            download_file_path = downloads_dir / download_filename
        
        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if isinstance(data[0], dict):
            # Data is list of dictionaries
            fieldnames = headers or list(data[0].keys())
            
            # Add headers with styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row_num, row_data in enumerate(data, 2):
                for col_num, fieldname in enumerate(fieldnames, 1):
                    ws.cell(row=row_num, column=col_num, value=row_data.get(fieldname, ""))
        
        else:
            # Data is list of lists
            if headers:
                # Add headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                start_row = 2
            else:
                start_row = 1
            
            # Add data rows
            for row_num, row_data in enumerate(data, start_row):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(str(final_file_path))
        
        # If we created a separate download copy, save that too
        if 'download_file_path' in locals():
            wb.save(str(download_file_path))
            download_url = f"http://localhost:8000/files/download/{download_filename}"
            public_filename = download_filename
        else:
            download_url = f"http://localhost:8000/files/download/{final_file_path.name}"
            public_filename = final_file_path.name
        
        return {
            "success": True,
            "file_path": str(final_file_path),
            "download_url": download_url,
            "public_filename": public_filename,
            "sheet_name": sheet_name,
            "rows_written": len(data),
            "columns": len(headers or (list(data[0].keys()) if isinstance(data[0], dict) else data[0])),
            "attachments": [
                {"type": "text", "content": f"✅ Created Excel file: {public_filename} with {len(data)} rows in sheet '{sheet_name}'\n\n📥 Download link: {download_url}"}
            ]
        }
    except Exception as e:
        return {"error": f"Error creating Excel file: {str(e)}"}